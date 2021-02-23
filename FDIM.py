# ---------------------------- GEN5 VRTT - FDIM ------------------------------
# ------------------------------------------------------------------------------------
#   Intel Confidential
# ------------------------------------------------------------------------------------
# - This script runs Frequency Domain Impedance Measurement
# - The load frequency is swept from 1KHz to 2.5MHz. This can be changed by altering Fr_List, F_Start and F_End.
# - Complex values of impedance @ each frequency point is outputted to an excel sheet
# - Impedance vs Frequency is plotted on excel using the data points
# ------------------------------------------------------------------------------------
#
#   Rev 1.0   - 5/15/2018
#           - Initial versioning.
#   pulled latest from GEN5CONTROLLER\NEXTGEN\PythonApplication2

import clr,sys,os,time,openpyxl,numpy
from math import log10
from openpyxl import load_workbook, Workbook
from openpyxl.chart import ScatterChart, Reference, Series

sys.path.append(os.path.join(sys.path[0],'intel'))

from Common.APIs import DataAPI,DisplayAPI,MeasurementAPI,GeneratorAPI,MeasurementItemsAPI
from Common.Models import TriggerModel,FanModel,RawDataModel,DataModel,DisplayModel,MeasurementItemModel
from Common.Enumerations import HorizontalScale,Cursors,ScoplessChannel,PowerState,Transition,Protocol,RailTestMode,MeasurementItemType,MeasurementItemUnit

from System import String, Char, Int32, UInt16, Boolean, Array, Byte, Double
from System.Collections.Generic import List,Dictionary


# Test Parameters - Edit these
test_rail = "VCCGT"
test_voltage = 0.9
high_current = 28  # High current level (A)
low_current = 18  # Low current level (A)
duty_cycle = 32  # Duty cycle (%)
start_frequency = 1  # Start frequency (kHz)
end_frequency = 1000  # End frequency (kHz)
drive_rise_time = 100
drive_fall_time = 100
samples_per_decade = 50

num_decades = log10(end_frequency) - log10(start_frequency)
frequency_list = numpy.logspace(log10(start_frequency), log10(end_frequency), int(num_decades * samples_per_decade), True)
# frequency_list = [10,20,30,123]  # Can specify frequencies explicitly


generator_api = GeneratorAPI()
rails = generator_api.GetRails()
data_api = DataAPI()
display_api = DisplayAPI()
measurement_api = MeasurementAPI()
measurement_items_api = MeasurementItemsAPI()


# Set-up Excel  #########
wb = Workbook()
ws1 = wb.create_sheet(title="Load Test", index=1)
file_name = 'FDIM' + time.strftime("_%j_%M_%S", time.localtime()) + ".xlsx"

# End Of imports
print("##############################################################################")
print("------------------------------------------------------------------------------")
print("                    FREQ DOMAIN IMPEDANCE MEASUREMENT                         ")
print("------------------------------------------------------------------------------")
print("##############################################################################")
print("")

# What Rails are available?
print("Displaying rails available")
print("------------------------------------------")
t = generator_api.GetRails()
for i in range(0, len(t)):
    print(t[i].Name, " Max Current - ", t[i].MaxCurrent, "; Vsense - ", t[i].VSenseName)
    print("Load sections: " + t[i].LoadSections)
    for k in t[i].ProtoParm.Keys:
        print("      ", k, " ", t[i].ProtoParm[k])
print("")
print("")
print("Begining Test")
print("--------------")

# Change Fan Speed
data_api.SetFanSpeed(10)

# Set trigger
data_api.SetTrigger(4, 0, 0, 0, 0, False)

# Enable display
display_api.Ch1_2Rail(test_rail)
display_api.SetChannel(ScoplessChannel.Ch1, True)
display_api.SetChannel(ScoplessChannel.Ch2, True)

# Assign rail to tab 1
generator_api.AssignRailToDriveOne(test_rail)   

# Set Voltage
generator_api.SetVoltageForRail(test_rail, test_voltage, Transition.Fast)

time.sleep(0.25)
measurement_api.MeasureCurrentFrequency(test_rail)


print("")
print("Freq(kHz) \t Impedance(Ohms)")
print("--------------------------------------------------------------------------------")
for i, frequency in enumerate(frequency_list):

    # Setup display
    horizontal_scale = 0
    # Set the horizontal scale such that at least 2 complete
    # Get minimum period (convert to microseconds)...
    period_us = (1 / (frequency * 1000)) * 1000000
    minimum_time = period_us * 4

    if minimum_time >= 5000 * 8 and horizontal_scale != HorizontalScale.Scale5ms:
        display_api.SetHorizontalScale(HorizontalScale.Scale5ms)
        horizontal_scale = HorizontalScale.Scale5ms
    elif minimum_time >= 2000 * 8 and horizontal_scale != HorizontalScale.Scale2ms:
        display_api.SetHorizontalScale(HorizontalScale.Scale2ms)
        horizontal_scale = HorizontalScale.Scale2ms
    elif minimum_time >= 1000 * 8 and horizontal_scale != HorizontalScale.Scale1ms:
        display_api.SetHorizontalScale(HorizontalScale.Scale1ms)
        horizontal_scale = HorizontalScale.Scale1ms
    elif minimum_time >= 500 * 8 and horizontal_scale != HorizontalScale.Scale500us:
        display_api.SetHorizontalScale(HorizontalScale.Scale500us)
        horizontal_scale = HorizontalScale.Scale500us
    elif minimum_time >= 200 * 8 and horizontal_scale != HorizontalScale.Scale200us:
        display_api.SetHorizontalScale(HorizontalScale.Scale200us)
        horizontal_scale = HorizontalScale.Scale200us
    elif minimum_time >= 80 * 8 and horizontal_scale != HorizontalScale.Scale80us:
        display_api.SetHorizontalScale(HorizontalScale.Scale80us)
        horizontal_scale = HorizontalScale.Scale80us
    elif minimum_time >= 40 * 8 and horizontal_scale != HorizontalScale.Scale40us:
        display_api.SetHorizontalScale(HorizontalScale.Scale40us)
        horizontal_scale = HorizontalScale.Scale40us
    elif minimum_time >= 20 * 8 and horizontal_scale != HorizontalScale.Scale20us:
        display_api.SetHorizontalScale(HorizontalScale.Scale20us)
        horizontal_scale = HorizontalScale.Scale20us
    elif horizontal_scale != HorizontalScale.Scale10us:
        display_api.SetHorizontalScale(HorizontalScale.Scale10us)
        horizontal_scale = HorizontalScale.Scale10us
    # display_api.SetHorizontalRange([0, horizontal_scale*8])

    # This comes in handy when converting from index to time
    time_per_point = horizontal_scale * 8 / 8192  # us
    sample_rate = 1 / (time_per_point / 1000000)
    # Set the current
    generator_api.Generator1DualSlopeSVDC(test_rail, high_current, low_current, frequency, duty_cycle, drive_rise_time, drive_fall_time, True)

    time.sleep(1)

    traces = data_api.GetVoltageCurrent(test_rail, -1)

    # Put the traces
    cur = numpy.array([100 * current for current in traces.Current])
    vol = numpy.array([100 * voltage for voltage in traces.Voltage])

    # Get fft
    cur_fft = numpy.fft.fft(cur, len(cur))
    vol_fft = numpy.fft.fft(vol, len(vol))
    frequency_measure = measurement_items_api.GetCurrentFrequencyOnce(test_rail, -1)[0].Value

    # Get the index of the bin that has our set frequency in it
    binsize = (sample_rate / 8192)  # Hz
    fft_idx = int(round((frequency * 1000) / binsize))

    impedance = numpy.absolute(vol_fft[fft_idx]) / numpy.absolute(cur_fft[fft_idx])
    phase = numpy.arctan(numpy.imag(vol_fft[fft_idx]) / numpy.real(vol_fft[fft_idx])) - numpy.arctan(numpy.imag(cur_fft[fft_idx]) / numpy.real(cur_fft[fft_idx]))
    if (phase > 0):
        phase = phase - 3.14

    voltage_amplitude = (max(vol) - min(vol)) / 100
    current_amplitude = (max(cur) - min(cur)) / 100

    print(round(frequency, 1), '\t \t', impedance)

# Store Data
    Row = 2 + i
    ws1.cell(column=4, row=Row, value=frequency_measure)
    ws1.cell(column=5, row=Row, value=frequency)
    ws1.cell(column=6, row=Row, value=impedance)
    ws1.cell(column=7, row=Row, value=phase)
    ws1.cell(column=8, row=Row, value=voltage_amplitude)
    ws1.cell(column=9, row=Row, value=current_amplitude)


# Set up Excel Doc:
ws1.cell(column=1, row=2, value="Test Rail")
ws1.cell(column=2, row=2, value=test_rail)
ws1.cell(column=1, row=3, value="High Current")
ws1.cell(column=2, row=3, value=high_current)
ws1.cell(column=1, row=4, value="Low Current")
ws1.cell(column=2, row=4, value=low_current)
ws1.cell(column=1, row=5, value="Duty Cycle")
ws1.cell(column=2, row=5, value=duty_cycle)
ws1.cell(column=1, row=6, value="Drive Rise Time")
ws1.cell(column=2, row=6, value=drive_rise_time)
ws1.cell(column=1, row=7, value="Drive Fall Time")
ws1.cell(column=2, row=7, value=drive_fall_time)


ws1.cell(column=4, row=1, value="Measured Frequency(kHz)")
ws1.cell(column=5, row=1, value="Set Frequency(kHz)")
ws1.cell(column=6, row=1, value="Impedance(Ohms)")
ws1.cell(column=7, row=1, value="Phase(Degrees)")
ws1.cell(column=8, row=1, value="Voltage Amplitude(V)")
ws1.cell(column=9, row=1, value="Current Amplitude(A)")


# Test done. Return to SVSC mode and switch Load off
generator_api.Generator1SVSC(test_rail, 0, False)

# save data to excel and plot Z vs F
wb.save(filename=file_name)
chart = ScatterChart()
chart.title = "FDIM"
chart.style = 6
chart.x_axis.title = 'Frequency (kHz)'
chart.y_axis.title = 'Impedance (Ohms)'

xvalues = Reference(ws1, min_col=5, min_row=2, max_row=Row)
values = Reference(ws1, min_col=6, min_row=2, max_row=Row)
series = Series(values, xvalues, title_from_data=False)
chart.series.append(series)
series.marker = openpyxl.chart.marker.Marker('circle')
series.graphicalProperties.line.noFill = True

ws1.add_chart(chart, "G2")

wb.save(filename=file_name)
print("--------------------------------------------------------------------------------")
print("")
print("Test complete. Results exported to Excel")
