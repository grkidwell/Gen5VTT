
# Rev 1.3   -7/27/2020
#           -bug fixed

# Rev 1.2   -1/9/2020
#           - Depopulated some vectors () arguments.
#
# Rev 1.1   -10/22/2019
#           - Depopulated some vectors () arguments.
#           - Added rail, current input from user

# Rev 1.0   - 8/13/2018
#           - Initial version
#             Changed LoadAllVectors function to use svid bus number

# ---------------------------- GEN5 VRTT - Imon Timing -------------------------------
# ------------------------------------------------------------------------------------
#   Intel Confidential 
# ------------------------------------------------------------------------------------
# - This script executes multiple (16) IMON reads in one half load cycle 
# - Load frequency can be changed by altering the parameter "Frequency" 
# - Response of each SVID transaction is displayed on Python shell
# - Twenty runs are executed in total
#   - Ten runs corresponding to Imin-Imax transition 
#   - Ten runs corresponding to Imax-Imin transition
# - IMON values of each run are ouputted to an Excel sheet
#   Rev 1.0   - 5/15/2018
#           - Initial versioning.
#   pulled latest from GEN5CONTROLLER\NEXTGEN\PythonApplication2
# ------------------------------------------------------------------------------------
import clr
import sys, os
import time

sys.path.append(os.path.join(sys.path[0],'intel'))

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
from openpyxl.drawing.image import Image

from Common.APIs import DataAPI
from Common.APIs import DisplayAPI
from Common.APIs import VectorAPI
from Common.APIs import MainAPI
from Common.Models import TriggerModel
from Common.Models import FanModel
from Common.Models import RawDataModel
from Common.Models import DataModel
from Common.Models import DisplayModel
from Common.Enumerations import HorizontalScale
from Common.Enumerations import Cursors
from Common.APIs import MeasurementAPI
from Common.APIs import GeneratorAPI
#from CommonBL import *
from Common.Enumerations import PowerState
from Common.Enumerations import Transition
from Common.Enumerations import PowerState
from Common.Enumerations import Transition
from Common.Enumerations import Protocol
from Common.Enumerations import RailTestMode
from Common.Enumerations import ScoplessChannel


from System import String, Char, Int32, UInt16, Boolean, Array, Byte, Double
from System.Collections.Generic import List
from System.Collections.Generic import Dictionary


gen = GeneratorAPI();
rails = gen.GetRails();
data = DataAPI()
display = DisplayAPI()
Measurment=MeasurementAPI()
vector = VectorAPI()
mn=MainAPI()


##Set-up Excel  #########
wb=Workbook()
ws1 = wb.create_sheet(title="Load Test", index=1)


from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

################################End Of imports

print("##############################################################################")
print("------------------------------------------------------------------------------")
print("                            IMON READ TEST                                    ")
print("------------------------------------------------------------------------------")
print("##############################################################################")
print("")

## What Rails are available?
print ("Displaying rails available");
print ("------------------------------------------"); 
t=gen.GetRails()
for i in range (0, len(t)):
    print(t[i].Name, " Max Current - ", t[i].MaxCurrent, "; Vsense - ", t[i].VSenseName)
    print("Load sections: " + t[i].LoadSections)
    for k in t[i].ProtoParm.Keys:
                print("      ", k, " ", t[i].ProtoParm[k])
print ("");
print (""); 
print ("Beginning Test");
print ("--------------");

##Change Fan Speed
data.SetFanSpeed(10)    

#Set Scale
display.SetHorizontalScale(HorizontalScale.Scale80us)
display.SetVerticalCurrentScale(0, 330)
display.SetVerticalVoltageScale(0, 2)

## Set trigger
data.SetTrigger(4, 0, 0, 0, 0, False)

#Specify Test Parameteres
#######################


#test_rail
voltage_rail=input('enter Rail Name(Refer rail name above): ')

#rail address
rail_address=input('enter rails address number:')

# MaxCurrent
I_Max =  input('enter max current:')
# MinCurrent
I_Min =input('enter min current:')

I_Max_i=I_Min
I_Min_i=I_Max


#####Assign rails to Tabs
gen.AssignRailToDriveOne(voltage_rail)

# DutyCycle
DC=50
# F start (kHz)
Frequncy=1.6666
#Ramp
Ramp=150

######################## Vector functions ##############################################
# Create vectors to read IMON, define delay between vectors and load into FPGA.  
def SetVectorsforMode4Int():
    t_svid=UInt16(1/(2*1000*Frequncy*0.00000004*15)-39)
    for i in range (0,16):
# Arguments in order -
# Index, FrameStart, VR Addr, SVID Cmd, SVID Payload, Parity (0-odd, 1-even, 2-Auto), FrameEnd, Delay, Status (1-active, 0-Inactive)
        vector.CreateVector(i,2,rail_address,7,0x15,2,3,t_svid,1)
    time.sleep(0.1)
    vector.LoadAllVectors(0,1)
    time.sleep(0.1)
    vector.VectorTriggerSettingsMode4(0,0)
    time.sleep(0.1)
    vector.ExecuteVectors(2)
    time.sleep(0.5)
    vector.ReadVectorResponses()
    vResult = data.GetVectorDataOnce()
    print('ACK\tDATA\tPRTYERR\tPRTYVAL')
    print('-----------------------------')
    for i in range (0,16):
        print(str(vResult.AckResp[i]),'\t',str(vResult.DataResp[i]),'\t',str(vResult.ParityErrResp[i]),'\t',str(vResult.ParityValResp[i]))
    return vResult.DataResp

# Repeat vectors execution. "SetVectorsMode4Int()" needs to precede this function.   
def RepeatVector(DACDelay):
    vector.RestartMode4()
    #vector.ReadVectorResponses(16)
    time.sleep(0.5)
    vResult = data.GetVectorDataOnce()
    print('ACK\tDATA\tPRTYERR\tPRTYVAL')
    print('-----------------------------')
    for i in range (0,16):
        print(str(vResult.AckResp[i]),'\t',str(vResult.DataResp[i]),'\t',str(vResult.ParityErrResp[i]),'\t',str(vResult.ParityValResp[i]))
    return vResult.DataResp


# Reset all vectors
def ResetAllVectors():
    vector.ResetVectors()

###### Display CSOn and Set Voltage 
display.DisplayDigitalSignal('CSO1#')
#gen.SetVoltageForRail(voltage_rail, voltage_level, Transition.Fast)

##Enable Voltage/Current mean
time.sleep(0.25)
display.Ch1_2Rail(voltage_rail)
display.SetChannel(ScoplessChannel.Ch1, True)
display.SetChannel(ScoplessChannel.Ch2, True)

Measurment.MeasureCurrentFrequency(voltage_rail)


#Create Excel File
ws1 = wb.create_sheet(title="IMON_Time", index=1)
file_name='IMON_Time' + time.strftime("_%j_%M_%S", time.localtime()) +".xlsx"

#Set up Excel Doc:

ws1.cell(column=1, row=3, value="High Current")
ws1.cell(column=1, row=4, value=I_Max)
ws1.cell(column=1, row=5, value="Low Current")
ws1.cell(column=1, row=6, value=I_Min)
ws1.cell(column=1, row=7, value="Duty Cycle")
ws1.cell(column=1, row=8, value=DC)
ws1.cell(column=1, row=9, value="Ramp")
ws1.cell(column=1, row=10, value=Ramp)

ws1.cell(column=1, row=11, value="Frequncy")
ws1.cell(column=1, row=12, value=Frequncy)




#Main Loop
gen.Generator1SVDC(voltage_rail, I_Max, I_Min, Frequncy, DC, Ramp, True)
####Enable Scope Traces
display.SetChannel(ScoplessChannel.Ch1, True)
display.SetChannel(ScoplessChannel.Ch2, True)


## Set trigger to "SVID Burst"
data.SetTrigger(12, 0, 0, 0, 0, False)
time.sleep(1)
# Set vectors and execute 
res = SetVectorsforMode4Int()
time.sleep(1)
mn.CaptureScopeless("IMON_low.png")
time.sleep(1)
img = Image('C:/Gen5/Images/IMON_low.png')
ws1.add_image(img, 'N2')
for i in range (1, len(res)+1):
    ws1.cell(column=3, row=i+1, value=res[i-1])

res=None


# calculate delay for 180 degree DAC phase shift     
t_half=int((0.5/Frequncy)/0.000008)
gen.Generator1SVDC(voltage_rail, I_Max_i, I_Min_i, Frequncy, DC, Ramp, True)
res = RepeatVector(t_half)
time.sleep(1)
mn.CaptureScopeless("IMON_high.png")
time.sleep(1)
img = Image('C:/Gen5/Images/IMON_High.png')
ws1.add_image(img, 'N31')
for i in range (1, len(res)+1):
    ws1.cell(column=3, row=31+i+1, value=res[i-1])
res=None    



for j in range(1, 10):
    gen.Generator1SVDC(voltage_rail, I_Max, I_Min, Frequncy, DC, Ramp, True)
    res=RepeatVector(0)
    time.sleep(0.5)
    for i in range (1, len(res)+1):
        ws1.cell(column=3+j, row=i+1, value=res[i-1])
    time.sleep(0.5)
    res=None
    gen.Generator1SVDC(voltage_rail, I_Max_i, I_Min_i, Frequncy, DC, Ramp, True)
    t_half=int((Frequncy/2)/0.000008)
    res=RepeatVector(t_half)
    time.sleep(0.5)
    for i in range (1, len(res)+1):
        ws1.cell(column=3+j, row=31+i+1, value=res[i-1])
    time.sleep(0.5)
    res=None


#All done - reset vectors 
ResetAllVectors()

#Switch off Load
gen.Generator1SVDC(voltage_rail, I_Max, I_Min, Frequncy, DC, Ramp, False)
data.SetTrigger(4, 0, 0, 0, 0, False)
time.sleep(5)

# Test done - Reset vectors and save excel 
ResetAllVectors()
wb.save(filename = file_name)
print ("--------------------------------------------------------------------------------")
print ("")
print ("Test complete. Results exported to Excel");
