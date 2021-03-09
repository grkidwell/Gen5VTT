# ------------------------------------------------------------------------------------
# ----Author - G. Kidwell
# 
# 
# ---Modified from Intel DC_Load_Example
# 
# NOTE: BEFORE RUNNING THIS PROGRAM, YOU MUST ATTACH SORENSEN MML-2 LOAD TO EVALBOARD AND TO PC VIA GPIB
#  AND LAUNCH THE GEN5VTT TOOL SOFTWARE
#
#--------------------------------------------------------------------------------------

import clr, os, sys, time
import numpy as np
from instr.gpib_equip import SorensonMM2

import openpyxl
from openpyxl import load_workbook,Workbook

#sys.path.append(os.path.join(sys.path[0],'intel'))

from Common.APIs import DataAPI, DisplayAPI, MeasurementAPI, MeasurementItemsAPI, GeneratorAPI,VectorAPI
from Common.Models import TriggerModel,FanModel,RawDataModel,DataModel,DisplayModel
from Common.Enumerations import HorizontalScale, ScoplessChannel, Cursors,PowerState,Transition,Protocol,RailTestMode

generator_api = GeneratorAPI();  rails = generator_api.GetRails();    data_api = DataAPI()
display_api = DisplayAPI();      measurement_api = MeasurementAPI();  measurement_items_api = MeasurementItemsAPI()
vector_api=VectorAPI()

def define_rail_data(testrail):
    rail = None
    for r in list(generator_api.GetRails()):
        if r.Name == testrail:
            rail = r
    if rail is None:
        print('Test failed because the test rail could not be found. Please check the rail name.')
    return rail

def setup_tool_display(testrail):
    generator_api.AssignRailToDriveOne(testrail)
    data_api.SetFanSpeed(10); display_api.SetHorizontalScale(HorizontalScale.Scale10us)
    display_api.SetVerticalCurrentScale(0,100);        display_api.SetVerticalVoltageScale(0,2)
    data_api.SetTrigger(4, 0, 0, 0, 0, False);         display_api.Ch1_2Rail(testrail, True) 
    display_api.SetChannel(ScoplessChannel.Ch1, True); display_api.SetChannel(ScoplessChannel.Ch2, True)

def setup_initial_voltage(testrail,testvoltage):
    generator_api.Generator1SVSC(testrail, 0, False); generator_api.SetVoltageForRail(testrail, testvoltage, Transition.Fast)

def enable_measurements(testrail):
    measurement_api.MeasureCurrentMean(testrail);      measurement_api.MeasureVoltageMean(testrail)
    measurement_api.MeasureVoltageAmplitude(testrail); measurement_api.MeasureVoltageMinMax(testrail)

def measure_vout_ripple(testrail):
    vout=measurement_items_api.GetVoltageMeanOnce(testrail)[0].Value
    vmin_vmax_measurement = measurement_items_api.GetVoltageMinMaxOnce(testrail, -1)
    vmax = vmin_vmax_measurement[0].Value; vmin = vmin_vmax_measurement[1].Value
    vpp = int((vmax - vmin)*1000)
    vpos_ripple=int((vmax-vout)*1000); vneg_ripple=int((vout-vmin)*1000)
    return vout,vpp,vpos_ripple,vneg_ripple

class Loadline:
    def __init__(self,raildata,testrail,powerstate,testcurrents,iccmax):
        self.load1 = SorensonMM2(active_ch=1); self.load2 = SorensonMM2(active_ch=3)
        self.load1max=25;                      
        self.raildata=raildata;                self.testrail=testrail
        self.ps=int(powerstate);               self.testcurrents=testcurrents
        self.iccmax=iccmax;                    self.voffset_cal = 0.001
    
    def scale_imon(self,imon_hex_value):
        return np.round(int(imon_hex_value,16)/0xff*self.iccmax,3)

#this function needs intel GUI svid tool to be open first so was replaced with config_imon_vector
#and read_imon_vector
    def read_imon(self):  
        raildata=self.raildata
        svid_transaction = None
        if raildata is not None:
            svid_command = 0x7; imon_register_address = 0x15
            data_api.SetSvidCmdWrite(raildata.VRAddress, svid_command, imon_register_address, raildata.SVIDBus)# raildata.SVIDBus,5,2,1,3)
            time.sleep(.25)
            svid_transaction = data_api.GetSvidData()
        svid_data=1
        if svid_transaction is not None:
            svid_data=svid_transaction.SVRData
        return svid_data

    def config_imon_vector(self):
        data_api.SetTrigger(4,0,0,0,0,False);time.sleep(0.1)
        vector_api.CreateVector(1,2,self.raildata.VRAddress,7,0x15,2,3,1,1);time.sleep(0.1)
        vector_api.LoadAllVectors(0,1);time.sleep(0.1)
        vector_api.VectorTriggerSettingsMode4(0,0);time.sleep(0.1)
        vector_api.ExecuteVectors(2);time.sleep(0.5)
        vector_api.ReadVectorResponses()
        
        
    def read_imon_vector(self):
        vector_api.RestartMode4();time.sleep(0.5)
        dimon = data_api.GetVectorDataOnce().DataResp[0]
        return f'{dimon:x}'


    def avg_dimon(self):
        return f'{int(sum([int(self.read_imon_vector(self.raildata),16) for count in range(3)])/3):x}' 

    def load(self,testcurrent):
        def meas_load(current,loadn):
            loadn.set_value(current); loadn.on(); time.sleep(1)
            return np.round(loadn.meas(),3)
        load2_pct=0.5*(testcurrent>self.load1max) #/100 
        total_load=meas_load(testcurrent*(1-load2_pct),self.load1)+ \
                   (meas_load(testcurrent*load2_pct,self.load2) if load2_pct else 0)
        return np.round(total_load,3)

    def set_ps(self):
        svid_command = 0x4
        data_api.SetSvidCmdWrite(rail_data.VRAddress, svid_command, self.ps, rail_data.SVIDBus)
        time.sleep(0.25)

    def take_data(self):
        self.set_ps()
        self.config_imon_vector()
        def get_values_and_print(current):
            load=self.load(current)
            vout,vpp,vpos_ripple,vneg_ripple = measure_vout_ripple(self.testrail)
            dd = {'load'  : load,            'vsense': vout+self.voffset_cal,
                  'ripple': vpp,             'vpos_ripple':vpos_ripple,
                  'vneg_ripple':vneg_ripple, 'dimon' : self.read_imon_vector()}
            dd['imon'] = self.scale_imon(dd['dimon'])
            print(f"{dd['load']}\t{dd['vsense']}\t{dd['ripple']}\t{dd['dimon']}\t{dd['imon']}")
            return dd
        self.dataset = [get_values_and_print(current) for current in self.testcurrents]
        self.load1.off(); self.load2.off(); self.load1.disconnect(); self.load2.disconnect()

    def export_to_excel(self):
        wb=Workbook(); ws1 = wb.create_sheet(title="loadline",index=1)
        file_name= os.path.join(sys.path[0],'Loadline' + time.strftime("_%j_%M_%S", time.localtime()) +".xlsx")
        start_row=3; start_col=1
        #column definitions by index for ps0-ps3
        col_dict = [{'1':'load','2':'vsense','3':'ripple','6':'dimon'},  
                    {'1':'load','2':'vsense','3':'ripple','6':'dimon'},  
                    {'1':'load','2':'vsense','3':'ripple','4':'vneg_ripple','5':'vpos_ripple','8':'dimon'}, 
                    {'1':'load','2':'vsense','3':'ripple','4':'vneg_ripple','5':'vpos_ripple','8':'dimon'}] 
        for col,label in col_dict[self.ps].items(): 
            ws1.cell(column=int(col)+start_col-1, row=start_row, value=label)
        for row, data_dict in enumerate(self.dataset):
            for col,label in col_dict[self.ps].items():
                ws1.cell(column=int(col)+start_col-1, row=row+start_row+1, value=data_dict[label])
        wb.save(filename = file_name)

#-----------------------------------------------------------------------------------------------
# this next section is WIP refactor of datastructures to streamline
# setting parameters based on whether Vcore or GT is selected and eventually power level


class Rail_params:
    def __init__(self):
        pass


rail_dict = {'VCCCORE':{'Iccmax':160, 'tdc':93},
             'VGT'    :{'Iccmax': 50, 'tdc':40}}

test_current_dict = {'PS0':'dummy'}

ps_dict = {'0':{'VID':0.9,'test_currents':'dummy'}}

#
#-------------------------------------------------------------------------------------------------------
# main program

test_rail = "VCCGT"
test_voltage = 0.3  # V
ps = 3
icc_max = 40
tdc = 20
start_current = 0.1
end_current = tdc
num_datapoints = 12
test_currents = [0,1]#np.linspace(start_current,end_current,num_datapoints)


rail_data = define_rail_data(test_rail)
print(rail_data.Name)

setup_tool_display(test_rail)
setup_initial_voltage(test_rail,test_voltage)
enable_measurements(test_rail)

loadline_dataset = Loadline(rail_data,test_rail,ps,test_currents,icc_max)


print("##################################")
print("----------------------------------")
print("          DC LOAD LINE            ")
print("----------------------------------")
print("##################################")
print("")
print("Begining Test")
print("")
print(f"If max load > {loadline_dataset.load1max}A connect 2nd load to evalboard")           
print("")
print('...............................')
print('load\tVsense\tRipple\tdIMON\tIMON')
print('...............................')

loadline_dataset.take_data()
vector_api.ResetVectors()
generator_api.Generator1SVSC(test_rail, 0, False)
data_api.SetTrigger(4, 0, 0, 0, 0, False)
loadline_dataset.export_to_excel()

print("-------------------------------")
print("")
print("Test complete.  Results exported to Excel")
